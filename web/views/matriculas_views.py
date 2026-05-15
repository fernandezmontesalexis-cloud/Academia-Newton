from django.shortcuts import render
from django.http import HttpResponse
from web.permisos import permiso_requerido
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from datetime import date

from ..models import Matricula
from ..utils import estado_ciclo_hoy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
@permiso_requerido(['admin', 'secretaria'])
def matriculas(request):
    sede = request.user.perfil.sede
    today = date.today()

    matriculas_qs = (
        Matricula.objects.filter(alumno__sede=sede, alumno__estado='activo')
        .select_related('alumno', 'alumno__sede', 'ciclo')
    )

    dni = request.GET.get('dni')
    if dni:
        matriculas_qs = matriculas_qs.filter(alumno__dni__icontains=dni)

    total_matriculas = matriculas_qs.count()
    activas = matriculas_qs.filter(
        ciclo__fecha_inicio__lte=today, ciclo__fecha_fin__gte=today
    ).count()
    finalizadas = matriculas_qs.filter(ciclo__fecha_fin__lt=today).count()
    pendientes = matriculas_qs.filter(ciclo__fecha_inicio__gt=today).count()

    paginator = Paginator(matriculas_qs.order_by('-fecha_matricula'), 10)
    matriculas_page = paginator.get_page(request.GET.get('page'))

    for m in matriculas_page:
        m.estado_academico = estado_ciclo_hoy(m.ciclo, today)

    return render(request, 'web/secretaria/matriculas/lista_matricula.html', {
        'matriculas': matriculas_page,
        'total_matriculas': total_matriculas,
        'activas': activas,
        'finalizadas': finalizadas,
        'pendientes': pendientes,
    })


@login_required
@permiso_requerido(['admin', 'secretaria'])
def exportar_excel(request):
    sede = request.user.perfil.sede
    today = date.today()

    matriculas_qs = (
        Matricula.objects.filter(alumno__sede=sede, alumno__estado='activo')
        .select_related('alumno', 'alumno__sede', 'ciclo')
        .order_by('-fecha_matricula')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrículas"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    headers = ['N°', 'Alumno', 'DNI', 'Ciclo', 'Fecha Matrícula', 'Sede', 'Estado Académico']
    for col, texto in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=texto)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, m in enumerate(matriculas_qs, 1):
        ws.append([
            i,
            str(m.alumno),
            m.alumno.dni,
            m.ciclo.nombre,
            m.fecha_matricula.strftime('%d/%m/%Y'),
            m.alumno.sede.nombre,
            estado_ciclo_hoy(m.ciclo, today).capitalize(),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="matriculas_{sede.nombre}.xlsx"'
    wb.save(response)
    return response
