import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
from decimal import Decimal

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from web.permisos import permiso_requerido
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, DecimalField, Value
from django.db.models.functions import Coalesce

from ..models import Sede, Alumno, Matricula, Pago, Ciclo


def _ingresos_6_meses(sede=None):
    """Calcula ingresos de los últimos 6 meses. Si sede es None, calcula globales."""
    hoy = date.today()
    labels = []
    data = []
    for i in range(5, -1, -1):
        mes = hoy.month - i
        año = hoy.year
        while mes <= 0:
            mes += 12
            año -= 1
        inicio = date(año, mes, 1)
        fin = date(año + 1, 1, 1) if mes == 12 else date(año, mes + 1, 1)
        qs = Pago.objects.filter(fecha_pago__gte=inicio, fecha_pago__lt=fin)
        if sede is not None:
            qs = qs.filter(matricula__alumno__sede=sede)
        total = qs.aggregate(total=Sum('monto'))['total'] or 0
        labels.append(inicio.strftime('%b %Y'))
        data.append(float(total))
    return labels, data


@login_required
@permiso_requerido(['admin'])
def reportes_sedes(request):
    data_sedes = []

    for sede in Sede.objects.all():
        alumnos_activos       = Alumno.objects.filter(sede=sede, estado='activo').count()
        matriculas_qs         = Matricula.objects.filter(alumno__sede=sede)
        matriculas_pagadas    = matriculas_qs.filter(estado='pagado').count()
        matriculas_pendientes = matriculas_qs.filter(estado='pendiente').count()

        ingresos = (
            Pago.objects.filter(matricula__alumno__sede=sede)
            .aggregate(total=Sum('monto'))['total'] or 0
        )
        deuda = sum(
            m.deuda() for m in
            matriculas_qs.filter(estado='pendiente')
            .select_related('ciclo').prefetch_related('pago_set')
        )

        # Datos gráfico 1: ingresos mensuales
        mensual_labels, mensual_data = _ingresos_6_meses(sede=sede)

        # Datos gráfico 3: esperado vs cobrado por ciclo
        ciclos = list(
            Matricula.objects.filter(alumno__sede=sede)
            .values('ciclo__id', 'ciclo__nombre', 'ciclo__precio')
            .annotate(total_mats=Count('id'))
            .order_by('ciclo__fecha_inicio')
        )
        bar_labels, bar_esperado, bar_cobrado = [], [], []
        for c in ciclos:
            bar_labels.append(c['ciclo__nombre'])
            esperado = float(c['ciclo__precio']) * c['total_mats']
            cobrado  = float(
                Pago.objects.filter(
                    matricula__ciclo_id=c['ciclo__id'],
                    matricula__alumno__sede=sede,
                ).aggregate(t=Sum('monto'))['t'] or 0
            )
            bar_esperado.append(round(esperado, 2))
            bar_cobrado.append(round(cobrado, 2))

        data_sedes.append({
            'sede':                   sede,
            'alumnos_activos':        alumnos_activos,
            'total_matriculas':       matriculas_qs.count(),
            'ingresos':               round(float(ingresos), 2),
            'deuda':                  round(float(deuda), 2),
            # Para gráficos modales
            'chart_mensual_labels':   json.dumps(mensual_labels),
            'chart_mensual_data':     json.dumps(mensual_data),
            'chart_estado_data':      json.dumps([matriculas_pagadas, matriculas_pendientes]),
            'chart_bar_labels':       json.dumps(bar_labels),
            'chart_bar_esperado':     json.dumps(bar_esperado),
            'chart_bar_cobrado':      json.dumps(bar_cobrado),
        })

    return render(request, 'web/administrador/reportes/reportes_sedes.html', {
        'data_sedes': data_sedes,
    })


@login_required
@permiso_requerido(['admin'])
def reporte_sede_detalle(request, sede_id):
    sede = get_object_or_404(Sede, id=sede_id)

    alumnos_activos = Alumno.objects.filter(sede=sede, estado='activo').count()
    matriculas_qs   = Matricula.objects.filter(alumno__sede=sede)

    ingresos_total = round(float(
        Pago.objects.filter(matricula__alumno__sede=sede)
        .aggregate(total=Sum('monto'))['total'] or 0
    ), 2)
    deuda_total = round(float(sum(
        m.deuda() for m in
        matriculas_qs.filter(estado='pendiente')
        .select_related('ciclo').prefetch_related('pago_set')
    )), 2)

    # Cards de ciclos — igual que secretaria en Reportes Financieros
    ciclos_raw  = Ciclo.objects.filter(sede=sede).order_by('-fecha_inicio')
    ciclos_data = []

    for c in ciclos_raw:
        total_matriculas = Matricula.objects.filter(
            ciclo=c, alumno__estado='activo', alumno__sede=sede,
        ).count()
        if total_matriculas == 0:
            continue

        total_recaudado = (
            Pago.objects.filter(
                matricula__ciclo=c,
                matricula__alumno__estado='activo',
                matricula__alumno__sede=sede,
            ).aggregate(t=Sum('monto'))['t'] or Decimal('0')
        )
        total_esperado = c.precio * total_matriculas
        total_deuda    = max(total_esperado - total_recaudado, Decimal('0'))

        alumnos_con_deuda = (
            Matricula.objects
            .filter(ciclo=c, alumno__estado='activo', alumno__sede=sede)
            .annotate(
                pagado_m=Coalesce(
                    Sum('pago__monto'), Value(Decimal('0')),
                    output_field=DecimalField(),
                )
            )
            .filter(pagado_m__lt=c.precio)
            .count()
        )

        pct = int(total_recaudado / total_esperado * 100) if total_esperado > 0 else 0

        ciclos_data.append({
            'ciclo':             c,
            'total_matriculas':  total_matriculas,
            'total_recaudado':   total_recaudado,
            'total_esperado':    total_esperado,
            'total_deuda':       total_deuda,
            'alumnos_con_deuda': alumnos_con_deuda,
            'pct_recaudado':     min(pct, 100),
        })

    return render(request, 'web/administrador/reportes/reporte_sede_detalle.html', {
        'sede':             sede,
        'alumnos_activos':  alumnos_activos,
        'total_matriculas': matriculas_qs.count(),
        'ingresos_total':   ingresos_total,
        'deuda_total':      deuda_total,
        'ciclos_data':      ciclos_data,
    })


@login_required
@permiso_requerido(['admin'])
def exportar_reporte_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center')

    headers = ['Alumno', 'DNI', 'Sede', 'Ciclo', 'Monto Pagado', 'Deuda', 'Fecha', 'Método']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    pagos = Pago.objects.select_related(
        'matricula__alumno',
        'matricula__alumno__sede',
        'matricula__ciclo',
    ).order_by('-fecha_pago')

    for row, pago in enumerate(pagos, 2):
        alumno = pago.matricula.alumno
        ws.cell(row=row, column=1, value=str(alumno))
        ws.cell(row=row, column=2, value=alumno.dni)
        ws.cell(row=row, column=3, value=str(alumno.sede))
        ws.cell(row=row, column=4, value=str(pago.matricula.ciclo))
        ws.cell(row=row, column=5, value=float(pago.monto))
        ws.cell(row=row, column=6, value=float(pago.matricula.deuda()))
        ws.cell(row=row, column=7, value=str(pago.fecha_pago))
        ws.cell(row=row, column=8, value=pago.get_metodo_pago_display())

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_financiero.xlsx"'
    wb.save(response)
    return response
