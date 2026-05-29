from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.urls import reverse

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import json

from web.permisos import permiso_requerido
from ..models import Matricula, Pago, Ciclo
from ..utils import estado_pago_matricula

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.colors import HexColor


@login_required
@permiso_requerido(['admin', 'secretaria'])
def pagos(request, matricula_id):
    try:
        matricula = Matricula.objects.select_related(
            'alumno__apoderado', 'alumno__sede', 'ciclo'
        ).get(id=matricula_id)
    except Matricula.DoesNotExist:
        return redirect(reverse('matriculas') + '?cancelada=1')

    if matricula.alumno.sede != request.perfil.sede:
        messages.error(request, "No tienes acceso a esta matrícula")
        return redirect('matriculas')

    apoderado = matricula.alumno.apoderado
    apoderado_nombre = apoderado.nombre_completo if apoderado else None

    def _contexto(error_monto=None, error_metodo=None, error_proximo=None,
                  monto_prev="", metodo_prev="", proximo_prev=""):
        total_pagado = Pago.objects.filter(matricula=matricula).aggregate(
            Sum('monto')
        )['monto__sum'] or Decimal('0')
        total_ciclo = matricula.ciclo.precio
        deuda = total_ciclo - total_pagado
        return {
            'matricula':        matricula,
            'total_pagado':     total_pagado,
            'total_ciclo':      total_ciclo,
            'deuda':            deuda,
            'apoderado_nombre': apoderado_nombre,
            'error_monto':      error_monto,
            'error_metodo':     error_metodo,
            'error_proximo':    error_proximo,
            'monto_prev':       monto_prev,
            'metodo_prev':      metodo_prev,
            'proximo_prev':     proximo_prev,
        }

    if request.method == 'POST':
        monto_raw        = request.POST.get('monto',        '').strip()
        metodo           = request.POST.get('metodo_pago',  '').strip()
        apo_value        = request.POST.get('apoderado',    '').strip()
        proximo_pago_raw = request.POST.get('proximo_pago', '').strip()

        total_pagado = Pago.objects.filter(matricula=matricula).aggregate(
            Sum('monto')
        )['monto__sum'] or Decimal('0')
        total_ciclo = matricula.ciclo.precio
        deuda_restante = total_ciclo - total_pagado

        error_monto   = None
        error_metodo  = None
        error_proximo = None

        try:
            monto = Decimal(monto_raw)
        except (InvalidOperation, TypeError, ValueError):
            error_monto = "Ingresa un monto numérico válido"
            monto = None

        if monto is not None:
            if monto <= 0:
                error_monto = "El monto debe ser mayor a 0"
            elif monto % Decimal('0.50') != 0:
                error_monto = "El monto debe ser en soles enteros o medios (ej. S/. 50.00 o S/. 50.50)"
            elif total_pagado + monto > total_ciclo:
                deuda_fmt = f"{deuda_restante:.2f}"
                error_monto = f"El monto excede la deuda restante (S/. {deuda_fmt})"

        if not metodo:
            error_metodo = "Debes seleccionar un método de pago"

        # Próximo pago obligatorio si el pago es parcial
        if monto is not None and error_monto is None:
            if monto < deuda_restante and not proximo_pago_raw:
                error_proximo = "Debe registrar una fecha de próximo pago para pagos parciales."
            elif proximo_pago_raw:
                try:
                    proximo_date = datetime.strptime(proximo_pago_raw, '%Y-%m-%d').date()
                    if proximo_date > matricula.ciclo.fecha_fin:
                        error_proximo = f"La fecha de próximo pago no puede superar el fin del ciclo ({matricula.ciclo.fecha_fin.strftime('%d/%m/%Y')})."
                except ValueError:
                    error_proximo = "Formato de fecha inválido."

        if error_monto or error_metodo or error_proximo:
            return render(
                request, 'web/secretaria/pagos/pagos.html',
                _contexto(error_monto=error_monto, error_metodo=error_metodo,
                          error_proximo=error_proximo, monto_prev=monto_raw,
                          metodo_prev=metodo, proximo_prev=proximo_pago_raw),
            )

        nuevo_pago = Pago.objects.create(
            matricula=matricula,
            registrado_por=request.perfil,
            fecha_pago=date.today(),
            monto=monto,
            metodo_pago=metodo,
            apoderado=apo_value,
        )

        # Pago.save() ya actualiza matricula.estado automáticamente.
        # Solo guardamos proximo_pago si cambió.
        if proximo_pago_raw:
            try:
                matricula.proximo_pago = datetime.strptime(proximo_pago_raw, '%Y-%m-%d').date()
            except ValueError:
                pass
        else:
            matricula.proximo_pago = None
        matricula.save()

        url = reverse('pagos', kwargs={'matricula_id': matricula.id})
        return redirect(f"{url}?ok=1&pago_id={nuevo_pago.id}")

    # Último pago registrado para esta matrícula (para el botón de boleta)
    ultimo_pago = Pago.objects.filter(matricula=matricula).order_by('-fecha_pago', '-id').first()

    return render(request, 'web/secretaria/pagos/pagos.html', {
        **_contexto(),
        'ultimo_pago': ultimo_pago,
    })


@login_required
@permiso_requerido(['admin', 'secretaria'])
def cancelar_matricula_nueva(request, matricula_id):
    """Elimina una matrícula solo si no tiene ningún pago registrado."""
    if request.method != 'POST':
        return redirect('matriculas')
    matricula = get_object_or_404(
        Matricula, id=matricula_id, alumno__sede=request.perfil.sede
    )
    total = Pago.objects.filter(matricula=matricula).aggregate(
        t=Sum('monto')
    )['t'] or 0
    if total == 0:
        alumno = matricula.alumno
        matricula.delete()
        if not Matricula.objects.filter(alumno=alumno).exists():
            alumno.delete()
    return redirect('matriculas')


def _matriculas_ciclo_qs(sede, ciclo):
    return (
        Matricula.objects
        .filter(alumno__sede=sede, alumno__estado='activo', ciclo=ciclo)
        .select_related('alumno', 'alumno__sede', 'ciclo')
        .prefetch_related('pago_set__registrado_por__user')
        .annotate(
            total_pagado_db=Coalesce(
                Sum('pago__monto'), Value(Decimal('0')),
                output_field=DecimalField(),
            )
        )
        .annotate(
            deuda_db=ExpressionWrapper(
                F('ciclo__precio') - F('total_pagado_db'),
                output_field=DecimalField(),
            )
        )
    )


def _pagos_json(m):
    pagos_data = []
    for p in m.pago_set.all().order_by('-fecha_pago'):
        pagos_data.append({
            'fecha':      p.fecha_pago.strftime('%d/%m/%Y'),
            'monto':      str(p.monto),
            'metodo':     p.get_metodo_pago_display(),
            'cajero':     p.registrado_por.user.username,
            'boleta_url': reverse('boleta_pdf', kwargs={'pago_id': p.id}),
        })
    return json.dumps(pagos_data, ensure_ascii=True)


@login_required
@permiso_requerido(['admin', 'secretaria'])
def lista_pagos(request):
    from django.core.paginator import Paginator

    sede  = request.perfil.sede
    today = date.today()

    ciclos_raw = Ciclo.objects.filter(sede=sede).order_by('-fecha_inicio')

    vigentes_data   = []   # activos, finalizan pronto y próximos
    finalizados_data = []  # ciclos ya cerrados

    for c in ciclos_raw:
        total_matriculas = Matricula.objects.filter(
            ciclo=c, alumno__estado='activo', alumno__sede=sede,
        ).count()

        if total_matriculas == 0:
            continue

        total_recaudado = (
            Pago.objects.filter(
                matricula__ciclo=c,
                matricula__alumno__estado='activo',
                matricula__alumno__sede=sede,
            ).aggregate(t=Sum('monto'))['t'] or Decimal('0')
        )

        total_esperado = c.precio * total_matriculas
        total_deuda    = max(total_esperado - total_recaudado, Decimal('0'))

        alumnos_con_deuda = (
            Matricula.objects
            .filter(ciclo=c, alumno__estado='activo', alumno__sede=sede)
            .annotate(
                pagado_m=Coalesce(
                    Sum('pago__monto'), Value(Decimal('0')),
                    output_field=DecimalField(),
                )
            )
            .filter(pagado_m__lt=c.precio)
            .count()
        )

        pct = int(total_recaudado / total_esperado * 100) if total_esperado > 0 else 0

        item = {
            'ciclo':             c,
            'total_matriculas':  total_matriculas,
            'total_recaudado':   total_recaudado,
            'total_esperado':    total_esperado,
            'total_deuda':       total_deuda,
            'alumnos_con_deuda': alumnos_con_deuda,
            'pct_recaudado':     min(pct, 100),
        }

        if c.alerta_vigencia == 'finalizado':
            finalizados_data.append(item)
        else:
            vigentes_data.append(item)

    # Paginación del historial: 6 tarjetas por página
    paginator       = Paginator(finalizados_data, 6)
    hpage           = request.GET.get('hpage', 1)
    finalizados_page = paginator.get_page(hpage)
    historial_abierto = 'hpage' in request.GET

    return render(request, 'web/secretaria/pagos/lista_pagos.html', {
        'vigentes_data':    vigentes_data,
        'finalizados_page': finalizados_page,
        'historial_abierto': historial_abierto,
        'total_finalizados': len(finalizados_data),
    })


@login_required
@permiso_requerido(['admin', 'secretaria'])
def reporte_ciclo(request, ciclo_id):
    sede  = request.perfil.sede
    today = date.today()

    ciclo = get_object_or_404(Ciclo, id=ciclo_id, sede=sede)

    mat_qs = _matriculas_ciclo_qs(sede, ciclo)

    estado_sel = request.GET.get('estado', '').strip()
    dni        = request.GET.get('dni',    '').strip()

    if dni:
        mat_qs = mat_qs.filter(alumno__dni__icontains=dni)
    if estado_sel == 'pagado':
        mat_qs = mat_qs.filter(deuda_db__lte=0)
    elif estado_sel == 'vencido':
        mat_qs = mat_qs.filter(deuda_db__gt=0, ciclo__fecha_fin__lt=today)
    elif estado_sel == 'parcial':
        mat_qs = mat_qs.filter(
            deuda_db__gt=0, total_pagado_db__gt=0, ciclo__fecha_fin__gte=today)
    elif estado_sel == 'sin_pago':
        mat_qs = mat_qs.filter(
            total_pagado_db=0, deuda_db__gt=0, ciclo__fecha_fin__gte=today)

    # Stat cards del ciclo
    pagos_hoy_c = (
        Pago.objects.filter(
            fecha_pago=today, matricula__ciclo=ciclo, matricula__alumno__sede=sede,
        ).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    )
    total_recaudado_c = (
        Pago.objects.filter(
            matricula__ciclo=ciclo, matricula__alumno__estado='activo',
            matricula__alumno__sede=sede,
        ).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    )
    total_matriculas_c = Matricula.objects.filter(
        ciclo=ciclo, alumno__estado='activo', alumno__sede=sede,
    ).count()
    total_deuda_c = max(
        ciclo.precio * total_matriculas_c - total_recaudado_c, Decimal('0')
    )
    alumnos_con_deuda_c = (
        Matricula.objects
        .filter(ciclo=ciclo, alumno__estado='activo', alumno__sede=sede)
        .annotate(
            pagado_m=Coalesce(
                Sum('pago__monto'), Value(Decimal('0')),
                output_field=DecimalField(),
            )
        )
        .filter(pagado_m__lt=ciclo.precio)
        .count()
    )

    from django.core.paginator import Paginator
    paginator = Paginator(
        mat_qs.order_by('alumno__apellido_paterno', 'alumno__nombres'), 12
    )
    matriculas_page = paginator.get_page(request.GET.get('page'))

    for m in matriculas_page:
        m.estado_pago = estado_pago_matricula(m, today)
        m.pagos_json  = _pagos_json(m)

    return render(request, 'web/secretaria/pagos/reporte_ciclo.html', {
        'ciclo':             ciclo,
        'matriculas':        matriculas_page,
        'pagos_hoy':         pagos_hoy_c,
        'total_recaudado':   total_recaudado_c,
        'total_deuda':       total_deuda_c,
        'alumnos_con_deuda': alumnos_con_deuda_c,
        'estado_sel':        estado_sel,
        'dni':               dni,
    })


@login_required
@permiso_requerido(['admin', 'secretaria'])
def exportar_reportes(request):
    sede = request.perfil.sede

    pagos_qs = (
        Pago.objects.filter(matricula__alumno__sede=sede)
        .select_related(
            'matricula__alumno', 'matricula__ciclo',
            'registrado_por__user',
        )
        .order_by('-fecha_pago', '-id')
    )

    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    metodo_sel  = request.GET.get('metodo', '').strip()
    ciclo_sel   = request.GET.get('ciclo', '').strip()
    dni         = request.GET.get('dni', '').strip()

    if fecha_desde:
        pagos_qs = pagos_qs.filter(fecha_pago__gte=fecha_desde)
    if fecha_hasta:
        pagos_qs = pagos_qs.filter(fecha_pago__lte=fecha_hasta)
    if metodo_sel:
        pagos_qs = pagos_qs.filter(metodo_pago=metodo_sel)
    if ciclo_sel:
        pagos_qs = pagos_qs.filter(matricula__ciclo_id=ciclo_sel)
    if dni:
        pagos_qs = pagos_qs.filter(matricula__alumno__dni__icontains=dni)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    headers = ['N°', 'Fecha', 'Alumno', 'DNI', 'Ciclo', 'Monto (S/.)', 'Método', 'Cajero']
    for col, texto in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=texto)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, p in enumerate(pagos_qs, 1):
        ws.append([
            i,
            p.fecha_pago.strftime('%d/%m/%Y'),
            str(p.matricula.alumno),
            p.matricula.alumno.dni,
            p.matricula.ciclo.nombre,
            float(p.monto),
            p.get_metodo_pago_display(),
            p.registrado_por.user.username,
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reportes_{sede.nombre}.xlsx"'
    wb.save(response)
    return response


@login_required
@permiso_requerido(['admin', 'secretaria'])
def boleta_pdf(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)

    if pago.matricula.alumno.sede != request.perfil.sede:
        return HttpResponse("No autorizado", status=403)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="comprobante_{pago.id:06d}.pdf"'

    width = 80 * mm
    height = 260 * mm

    doc = SimpleDocTemplate(
        response,
        pagesize=(width, height),
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=6 * mm,
        bottomMargin=6 * mm,
    )

    # Totales acumulados a la fecha del comprobante
    total_pagado = Pago.objects.filter(
        matricula=pago.matricula
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    deuda = pago.matricula.ciclo.precio - total_pagado

    alumno = pago.matricula.alumno
    ciclo = pago.matricula.ciclo

    # Paleta en escala de grises — apta para impresión B&W sin costo extra de tinta
    negro      = HexColor('#000000')
    gris_datos = HexColor('#1a1a1a')   # valores y datos principales
    gris_label = HexColor('#555555')   # etiquetas secundarias
    gris_sep   = HexColor('#bbbbbb')   # líneas separadoras

    # Estilos de párrafo
    titulo = ParagraphStyle(
        'titulo', fontName='Helvetica-Bold', fontSize=13,
        alignment=TA_CENTER, textColor=negro, spaceAfter=1,
    )
    subtitulo = ParagraphStyle(
        'subtitulo', fontName='Helvetica', fontSize=7,
        alignment=TA_CENTER, textColor=gris_label, spaceAfter=0,
    )
    numero_comp = ParagraphStyle(
        'numero_comp', fontName='Helvetica-Bold', fontSize=9,
        alignment=TA_CENTER, textColor=gris_datos, spaceAfter=0,
    )
    label = ParagraphStyle(
        'label', fontName='Helvetica', fontSize=6.5,
        textColor=gris_label, spaceAfter=1,
    )
    valor = ParagraphStyle(
        'valor', fontName='Helvetica-Bold', fontSize=8,
        textColor=gris_datos, spaceAfter=4,
    )
    monto_grande = ParagraphStyle(
        'monto_grande', fontName='Helvetica-Bold', fontSize=14,
        alignment=TA_CENTER, textColor=negro, spaceAfter=1,
    )
    footer_style = ParagraphStyle(
        'footer', fontName='Helvetica', fontSize=7,
        alignment=TA_CENTER, textColor=gris_label, spaceAfter=2,
    )

    def sep():
        return HRFlowable(
            width='100%', thickness=0.5,
            color=gris_sep, spaceAfter=4, spaceBefore=4,
        )

    # Estado del pago — solo texto, sin color (el texto ya lo comunica)
    if deuda <= 0:
        estado_texto = "PAGADO COMPLETAMENTE"
    elif total_pagado > pago.monto:
        estado_texto = "PAGO PARCIAL"
    else:
        estado_texto = "PENDIENTE DE PAGO"

    estado_style = ParagraphStyle(
        'estado', fontName='Helvetica-Bold', fontSize=9,
        alignment=TA_CENTER, textColor=negro, spaceAfter=2,
    )

    fecha_str = pago.fecha_pago.strftime('%d/%m/%Y')

    elementos = []

    # ── ENCABEZADO ──────────────────────────────────
    elementos.append(Spacer(1, 2 * mm))
    elementos.append(Paragraph("ACADEMIA NEWTON", titulo))
    elementos.append(Paragraph("Newton en Red  —  Sistema Académico", subtitulo))
    elementos.append(Spacer(1, 3 * mm))
    elementos.append(sep())

    # ── SEDE / FECHA ─────────────────────────────────
    elementos.append(Paragraph(f"Sede: {alumno.sede.nombre}", label))
    elementos.append(Paragraph(f"Fecha: {fecha_str}", label))
    elementos.append(sep())

    # ── NÚMERO DE COMPROBANTE ────────────────────────
    elementos.append(Spacer(1, 1 * mm))
    elementos.append(Paragraph(f"COMPROBANTE  N°  {pago.id:06d}", numero_comp))
    elementos.append(Spacer(1, 2 * mm))
    elementos.append(sep())

    # ── DATOS DEL ALUMNO ─────────────────────────────
    elementos.append(Paragraph("Alumno", label))
    elementos.append(Paragraph(str(alumno), valor))

    elementos.append(Paragraph("DNI", label))
    elementos.append(Paragraph(alumno.dni, valor))

    elementos.append(Paragraph("Ciclo", label))
    elementos.append(Paragraph(ciclo.nombre, valor))

    elementos.append(Paragraph("Cajero", label))
    elementos.append(Paragraph(pago.registrado_por.user.username, valor))

    if pago.apoderado:
        elementos.append(Paragraph("Apoderado / Responsable", label))
        elementos.append(Paragraph(pago.apoderado, valor))

    elementos.append(sep())

    # ── MÉTODO DE PAGO ───────────────────────────────
    elementos.append(Paragraph("Método de pago", label))
    elementos.append(Paragraph(pago.get_metodo_pago_display(), valor))
    elementos.append(sep())

    # ── MONTO DE ESTE PAGO ───────────────────────────
    elementos.append(Spacer(1, 2 * mm))
    elementos.append(Paragraph(f"S/. {pago.monto:.2f}", monto_grande))
    elementos.append(Paragraph("Monto de este pago", subtitulo))
    elementos.append(Spacer(1, 3 * mm))
    elementos.append(sep())

    # ── RESUMEN FINANCIERO ───────────────────────────
    elementos.append(Paragraph(f"Total del ciclo:      S/. {ciclo.precio:.2f}", label))
    elementos.append(Paragraph(f"Total pagado:         S/. {total_pagado:.2f}", label))

    deuda_style = ParagraphStyle(
        'deuda_line', fontName='Helvetica-Bold', fontSize=7,
        textColor=negro, spaceAfter=1,
    )
    elementos.append(Paragraph(
        f"Deuda restante:       S/. {deuda:.2f}",
        deuda_style,
    ))
    elementos.append(sep())

    # ── ESTADO ───────────────────────────────────────
    elementos.append(Spacer(1, 1 * mm))
    elementos.append(Paragraph(estado_texto, estado_style))
    elementos.append(sep())

    # ── FOOTER ───────────────────────────────────────
    elementos.append(Spacer(1, 3 * mm))
    elementos.append(Paragraph("Gracias por confiar en Academia Newton", footer_style))
    elementos.append(Paragraph("Newton en Red — Sistema Académico", footer_style))
    elementos.append(Spacer(1, 2 * mm))

    doc.build(elementos)
    return response
