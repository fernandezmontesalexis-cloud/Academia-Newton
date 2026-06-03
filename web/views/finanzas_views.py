import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from decimal import Decimal

from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.core.paginator import Paginator

from web.permisos import permiso_requerido
from ..models import Alumno, Matricula, Pago, Sede


_METODO_MAP = {
    'efectivo':      'Efectivo',
    'yape':          'Yape',
    'transferencia': 'Transferencia',
}


def _build_page_range(page_obj, window=2):
    """
    Genera los números de página para la paginación con puntos suspensivos.
    Por ejemplo: 1 ... 4 5 6 ... 20
    El parámetro window controla cuántas páginas mostrar alrededor de la actual.
    """
    total = page_obj.paginator.num_pages
    if total <= 7:
        return list(range(1, total + 1))
    current = page_obj.number
    result, last_added = [], None
    for i in range(1, total + 1):
        if i == 1 or i == total or abs(i - current) <= window:
            if last_added is not None and i - last_added > 1:
                result.append(None)  # None representa los "..."
            result.append(i)
            last_added = i
    return result


def _filtrar_historial(request):
    """
    Aplica los filtros del formulario de búsqueda al historial de pagos.
    Esta función la reutilizo tanto en la vista principal como en el exportador de Excel.
    """
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    sede_id         = request.GET.get('sede_id', '')
    metodo          = request.GET.get('metodo', '')
    busqueda        = request.GET.get('q', '').strip()

    # Traigo todos los pagos con sus relaciones — evita consultas extra al iterar
    qs = Pago.objects.select_related(
        'matricula__alumno',
        'matricula__alumno__sede',
        'matricula__ciclo',
    )

    # Aplico cada filtro solo si el usuario lo seleccionó
    if fecha_desde_str:
        try:
            qs = qs.filter(fecha_pago__gte=date.fromisoformat(fecha_desde_str))
        except ValueError:
            pass
    if fecha_hasta_str:
        try:
            qs = qs.filter(fecha_pago__lte=date.fromisoformat(fecha_hasta_str))
        except ValueError:
            pass
    if sede_id:
        qs = qs.filter(matricula__alumno__sede_id=sede_id)
    if metodo:
        qs = qs.filter(metodo_pago=metodo)
    if busqueda:
        # Busco por nombre o apellido del alumno
        qs = (
            qs.filter(matricula__alumno__nombres__icontains=busqueda) |
            qs.filter(matricula__alumno__apellido_paterno__icontains=busqueda)
        )

    # Armo la URL base para la paginación — preserva los filtros activos al cambiar de página
    params = request.GET.copy()
    params.pop('page', None)
    qs_str = params.urlencode()

    filtros = {
        'fecha_desde_str': fecha_desde_str,
        'fecha_hasta_str': fecha_hasta_str,
        'sede_id':         sede_id,
        'metodo':          metodo,
        'busqueda':        busqueda,
        'filtros_activos': bool(fecha_desde_str or fecha_hasta_str or sede_id or metodo or busqueda),
        'base_url':        f'?{qs_str}&' if qs_str else '?',
    }
    return qs.order_by('-fecha_pago', '-id'), filtros


@login_required
@permiso_requerido(['admin'])
def finanzas(request):
    hoy        = date.today()
    sem_inicio = hoy - timedelta(days=6)  # últimos 7 días incluyendo hoy

    # Traigo el historial ya filtrado — es la base de todos los cálculos de esta vista
    historial_qs, filtros = _filtrar_historial(request)
    historial = Paginator(historial_qs, 10).get_page(request.GET.get('page'))

    # ── KPI: Ingresos del período ─────────────────────────────────────────
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')

    if fecha_desde_str or fecha_hasta_str:
        # Si hay filtro de fecha, muestro el total del período filtrado
        ingresos_periodo = round(float(
            historial_qs.aggregate(t=Sum('monto'))['t'] or 0
        ), 2)
        periodo_label = "Total del período"
    else:
        # Sin filtro de fecha, muestro los ingresos del mes actual
        mes_qs = Pago.objects.filter(fecha_pago__gte=hoy.replace(day=1))
        if filtros['sede_id']:
            mes_qs = mes_qs.filter(matricula__alumno__sede_id=filtros['sede_id'])
        if filtros['metodo']:
            mes_qs = mes_qs.filter(metodo_pago=filtros['metodo'])
        ingresos_periodo = round(float(
            mes_qs.aggregate(t=Sum('monto'))['t'] or 0
        ), 2)
        periodo_label = "Ingresos del mes"

    # ── KPI: Deuda total pendiente ────────────────────────────────────────
    # La deuda respeta el filtro de sede pero no el de fecha — siempre muestra la deuda actual
    _mats_pend = Matricula.objects.filter(alumno__estado='activo', estado='pendiente')
    if filtros['sede_id']:
        _mats_pend = _mats_pend.filter(alumno__sede_id=filtros['sede_id'])
    _precio_pend = _mats_pend.aggregate(t=Sum('ciclo__precio'))['t'] or Decimal('0')
    _pagado_pend = (
        Pago.objects.filter(matricula__in=_mats_pend)
        .aggregate(t=Sum('monto'))['t'] or Decimal('0')
    )
    deuda_total = round(float(_precio_pend - _pagado_pend), 2)

    # ── KPI: Total cobrado esta semana ────────────────────────────────────
    sem_qs = Pago.objects.filter(fecha_pago__gte=sem_inicio)
    if filtros['sede_id']:
        sem_qs = sem_qs.filter(matricula__alumno__sede_id=filtros['sede_id'])
    if filtros['metodo']:
        sem_qs = sem_qs.filter(metodo_pago=filtros['metodo'])
    total_semana = round(float(
        sem_qs.aggregate(t=Sum('monto'))['t'] or 0
    ), 2)

    # Detecto el método de pago más usado en el historial filtrado
    metodo_result = (
        historial_qs.values('metodo_pago')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
        .first()
    )
    metodo_mas_usado = metodo_result['metodo_pago'] if metodo_result else ''
    metodo_label     = _METODO_MAP.get(metodo_mas_usado, '—')

    return render(request, 'web/administrador/finanzas/finanzas.html', {
        'ingresos_periodo': ingresos_periodo,
        'periodo_label':    periodo_label,
        'deuda_total':      deuda_total,
        'total_semana':     total_semana,
        'metodo_mas_usado': metodo_mas_usado,
        'metodo_label':     metodo_label,
        'historial':        historial,
        'page_range':       _build_page_range(historial),
        'sedes':            list(Sede.objects.all()),
        **filtros,
    })


@login_required
@permiso_requerido(['admin'])
def exportar_finanzas_excel(request):
    # Reutilizo los mismos filtros activos en pantalla para exportar exactamente lo que se ve
    historial_qs, _ = _filtrar_historial(request)
    pagos = historial_qs.select_related('registrado_por__user')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Financieros"

    # ── Estilos de la cabecera ────────────────────────────────────────────
    # Verde oscuro para diferenciarlo del Excel de reportes (que es azul)
    hdr_fill = PatternFill("solid", fgColor="1B5E20")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center   = Alignment(horizontal='center', vertical='center')
    left_al  = Alignment(horizontal='left',   vertical='center')

    cols = [
        ('N°',             6),
        ('Alumno',        28),
        ('DNI',           12),
        ('Sede',          16),
        ('Ciclo',         22),
        ('Monto Pagado',  14),
        ('Método',        14),
        ('Fecha',         13),
        ('Estado',        12),
        ('Deuda Restante',14),
        ('Registrado por',18),
    ]
    ws.row_dimensions[1].height = 22
    for col, (title, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Estilos de las filas de datos ─────────────────────────────────────
    fill_alt   = PatternFill("solid", fgColor="F5F5F5")   # gris claro para filas pares
    fill_comp  = PatternFill("solid", fgColor="E8F5E9")   # verde claro para pagos completos
    fill_parc  = PatternFill("solid", fgColor="FFF3E0")   # naranja claro para pagos parciales
    font_base  = Font(size=9)
    font_money = Font(size=9, color="1B5E20", bold=True)  # verde para montos pagados
    font_deuda = Font(size=9, color="B71C1C", bold=True)  # rojo para deuda pendiente
    font_parc  = Font(size=9, color="E65100", bold=True)
    curr_fmt   = '"S/. "#,##0.00'

    for row_num, pago in enumerate(pagos, 2):
        alumno    = pago.matricula.alumno
        matricula = pago.matricula
        estado    = 'Completo' if matricula.estado == 'pagado' else 'Parcial'
        deuda     = float(matricula.deuda())
        usuario   = pago.registrado_por.user.username if pago.registrado_por else '—'
        is_even   = (row_num % 2 == 0)
        ws.row_dimensions[row_num].height = 15

        filas_data = [
            (row_num - 1,                          center,  font_base,  None),
            (str(alumno),                          left_al, font_base,  None),
            (alumno.dni,                           center,  font_base,  None),
            (str(alumno.sede),                     left_al, font_base,  None),
            (matricula.ciclo.nombre,               left_al, font_base,  None),
            (float(pago.monto),                    center,  font_money, curr_fmt),
            (pago.get_metodo_pago_display(),       center,  font_base,  None),
            (pago.fecha_pago.strftime('%d/%m/%Y'), center,  font_base,  None),
            (estado,                               center,  font_base,  None),
            (deuda,                                center,  font_deuda if deuda > 0 else font_base, curr_fmt),
            (usuario,                              center,  font_base,  None),
        ]

        # La columna de Estado se colorea según si el pago está completo o parcial
        estado_fill = fill_comp if estado == 'Completo' else fill_parc

        for col, (value, align, font, fmt) in enumerate(filas_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font      = font
            cell.alignment = align
            if col == 9:
                # Columna Estado: color especial según completo/parcial
                cell.fill = estado_fill
                if estado == 'Parcial':
                    cell.font = font_parc
            elif is_even:
                # Filas pares: fondo gris suave para facilitar la lectura
                cell.fill = fill_alt
            if fmt:
                cell.number_format = fmt

    # Congelo la primera fila para que los encabezados siempre estén visibles al hacer scroll
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="finanzas_movimientos.xlsx"'
    wb.save(response)
    return response


@login_required
@permiso_requerido(['admin'])
def historial_completo(request):
    # Vista separada para el historial completo con los mismos filtros de finanzas
    historial_qs, filtros = _filtrar_historial(request)
    historial = Paginator(historial_qs, 10).get_page(request.GET.get('page'))

    return render(request, 'web/administrador/finanzas/historial_completo.html', {
        'historial':  historial,
        'page_range': _build_page_range(historial),
        'sedes':      list(Sede.objects.all()),
        **filtros,
    })
